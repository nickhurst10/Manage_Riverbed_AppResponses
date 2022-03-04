"""
This is a script to get configuration data from multiple Riverbed AppResponse 11 (AR11) and present them in a spreadsheet. The script requires two other libraries to work, “appresponse_device.py” and “appresponse_mgmt_api.py”, both of which are part of the github repository “https://github.com/nickhurst10/Manage_Riverbed_AppResponses.git”.
“appresponse_mgmt_api.py” interacts with the AR11 to do the restAPI GET requests.
“appresponse_device.py” manages the interaction between the main script and “appresponse_mgmt_api.py” library.

For the script to know what AR11 to access, this script looks for a file called “ar_list.csv” and looks for IP addresses under the header “ar_list”. An example is include in the github repository.

To run the script, the user must provide user credentials, which have the relevant access to all the AR11’s.

To this script was tested on python version 3.9.

Example to run the script
	python3 Manage_ARs_RO.py -u username

After which the user will be prompted to enter their password.

If you have any questions, please reach out to me.
"""
__author__ = "Nick Hurst nhurst@riverbed.com"


import requests
import json
from getpass import getpass
from datetime import datetime
import openpyxl
import appresponse_device
import argparse
import csv


COLUMN_TITLE_NUMBER = 1
MGMT_SPREADSHEET_PATH = 'AR_Management.xlsx'
AR_LIST_COL_NAME = 'AR_List'
API_CONFIG_FILE_NAME = 'api_config_info.conf'
AR_LIST_CSV_FILE_PATH = 'ar_list.csv'


AR_LIST_WORKSHEET  = 'AR_List'

NTP_WORKSHEET_NAME = 'ntp_information'
SNMP_WORKSHEET_NAME = 'snmp_info'
COMMOM_WORKSHEET_NAME = "common_info"
VIFGS_WORKSHEET_NAME = "vifgs_info"
CAP_JOBS_WORKSHEET_NAME = "cap_jobs_info"
DNS_WORKSHEET_NAME = "dns_info"
PHY_INT_WORKSHEET_NAME = "phy_int_info"
HOSTGROUPS_WORKSHEET_NAME = "hostgroup_info"
URL_WORKSHEET_NAME = "url_info"
APPLICATIONS_WORKSHEET_NAME = "apps_info"


NTP_COL_NAME_LIST = ["server_id","address","prefer","version","encryption","key_id","secret"]
SNMP_COL_NAME_LIST = ["enabled","contact","location","description","version","community_string","username","security_model","auth_protocol","authentication_passphrase","privacy_protocol","privacy_passphrase"]
COMMON_COL_NAME_LIST = ["device_name","sw_version","hw_version","mgmt_addresses","serial","model"]
VIFGS_COL_NAME_LIST = ["name","id","enabled","dedup","interfaces","description","filer_type","filter_value","bandwidth_capacity","is_other_vifg"]
CAP_JOBS_COL_NAME_LIST = ["name","vifgs","capture_from_all_vifgs","snap_len","enabled","microflow_index_enable","min_disk_space","max_disk_space","min_retention_time","max_retention_time","packet_data_optimize_for_read","min_disk_space","max_disk_space","min_retention_time","max_retention_time"]
DNS_COL_NAME_LIST = ["hostname","dns_servers","dns_domains"]
PHY_INT_COL_NAME_LIST = ["name","enabled","description","speed_duplex","interface_type","status","mtu"]
HOSTGROUPS_COL_NAME_LIST = ["name","description","enabled","hosts"]
URL_COL_NAME_LIST = ["name","description","enabled","preferred","urls"]
APPLICATIONS_COL_NAME_LIST = ["name","enable","traffic_match_mode","include_dpi_tags","hosts","ports","ip_protocol"]

def setup_worksheet(received_work_book,received_worksheet_name,received_columns_names_list):
    if received_worksheet_name in received_work_book.sheetnames:
        worksheet = received_work_book[received_worksheet_name]
        received_work_book.remove(worksheet)
    
    received_work_book.create_sheet(received_worksheet_name)
    worksheet = received_work_book[received_worksheet_name]

    worksheet.cell(1,1).value = "AR_IP_addr"
    col_num = 2
    for col_name in received_columns_names_list:
        worksheet.cell(1,col_num).value = col_name
        col_num += 1

def setup_management_spreadsheet(received_spreadsheet_path):

    ar_config_work_book = openpyxl.Workbook()

    setup_worksheet(ar_config_work_book,NTP_WORKSHEET_NAME,NTP_COL_NAME_LIST)
    setup_worksheet(ar_config_work_book,SNMP_WORKSHEET_NAME,SNMP_COL_NAME_LIST)
    setup_worksheet(ar_config_work_book,COMMOM_WORKSHEET_NAME,COMMON_COL_NAME_LIST)
    setup_worksheet(ar_config_work_book,VIFGS_WORKSHEET_NAME,VIFGS_COL_NAME_LIST)
    setup_worksheet(ar_config_work_book,CAP_JOBS_WORKSHEET_NAME,CAP_JOBS_COL_NAME_LIST)
    setup_worksheet(ar_config_work_book,DNS_WORKSHEET_NAME,DNS_COL_NAME_LIST)
    setup_worksheet(ar_config_work_book,PHY_INT_WORKSHEET_NAME,PHY_INT_COL_NAME_LIST)
    setup_worksheet(ar_config_work_book,HOSTGROUPS_WORKSHEET_NAME,HOSTGROUPS_COL_NAME_LIST)
    setup_worksheet(ar_config_work_book,APPLICATIONS_WORKSHEET_NAME,APPLICATIONS_COL_NAME_LIST)
    setup_worksheet(ar_config_work_book,URL_WORKSHEET_NAME,URL_COL_NAME_LIST)

    ar_config_work_book.save(received_spreadsheet_path)
    
def update_spreadsheet_with_config(received_spreadsheet_path,received_ar_config):

    ar_config_work_book = openpyxl.load_workbook(received_spreadsheet_path)
    #fill all relevant worksheets with the there relevant AR information
    update_ntp_worksheet(received_ar_config.ip_addr,received_ar_config.ntp_data,ar_config_work_book,NTP_WORKSHEET_NAME)
    update_snmp_worksheet(received_ar_config.ip_addr,received_ar_config.snmp_data,ar_config_work_book,SNMP_WORKSHEET_NAME)
    update_vifgs_worksheet(received_ar_config.ip_addr,received_ar_config.vifgs_data,ar_config_work_book,VIFGS_WORKSHEET_NAME)
    update_common_worksheet(received_ar_config.ip_addr,received_ar_config.common_data,ar_config_work_book,COMMOM_WORKSHEET_NAME)
    update_dns_worksheet(received_ar_config.ip_addr,received_ar_config.dns_data,ar_config_work_book,DNS_WORKSHEET_NAME)
    update_cap_jobs_worksheet(received_ar_config.ip_addr,received_ar_config.cap_job_data,ar_config_work_book,CAP_JOBS_WORKSHEET_NAME)
    update_phy_int_worksheet(received_ar_config.ip_addr,received_ar_config.phy_int_data,ar_config_work_book,PHY_INT_WORKSHEET_NAME)
    update_apps_worksheet(received_ar_config.ip_addr,received_ar_config.ar_apps_data,ar_config_work_book,APPLICATIONS_WORKSHEET_NAME)
    update_hostgroups_worksheet(received_ar_config.ip_addr,received_ar_config.hostgroups_data,ar_config_work_book,HOSTGROUPS_WORKSHEET_NAME)
    update_urls_worksheet(received_ar_config.ip_addr,received_ar_config.urls_data,ar_config_work_book,URL_WORKSHEET_NAME)
    ar_config_work_book.save(received_spreadsheet_path)

def find_column_with_title(received_title_name,received_workbook):
    for col_num in range(COLUMN_TITLE_NUMBER,received_workbook.max_column+1):
        if received_workbook.cell(1,col_num).value == received_title_name:
            return col_num

def get_ar_bearer_token(received_username,received_password,received_ar_ip_addr):
    #will get bearer token, will either return bearer token to a blank string to mean reat api access wasn't successful
    bearer_token=""
    url = f"https://{received_ar_ip_addr}/api/mgmt.aaa/1.0/token"
    payload = json.dumps({
        "user_credentials": {
            "username": received_username,
            "password": received_password
        }
        })
    headers = {
        'Content-Type': 'application/json'
        }
    try:
        response = requests.request("POST", url, headers=headers, data=payload,verify=False)
    except Exception as error_message:
        print (error_message)
    else:
        try:
            json_response = json.loads(response.text)
            print(f"bearer Token is: {json_response['access_token']}")
            bearer_token = json_response['access_token']
        except:
            print(f'failed to get bearer token from AR {received_ar_ip_addr} - data received was - \n\t{response.text}')
    return bearer_token

def confirm_rest_api_access_to_ARs_and_get_bearer_token(received_username,received_password,received_ar_ip_addr_list):
    ar_ip_addr_and_bearer_token_list = []
    for ar_ip_addr in received_ar_ip_addr_list:
        print(ar_ip_addr)
        bearer_token = (get_ar_bearer_token(received_username,received_password,ar_ip_addr))
        #if the received bearer token isn't blank then we know we can access the REST API of the AR so we add to the AR list with bearer token
        if bearer_token !="":
            ar_ip_addr_and_bearer_token={'bear_token':bearer_token, 'ar_ip_addr':ar_ip_addr}
            ar_ip_addr_and_bearer_token_list.append(ar_ip_addr_and_bearer_token)
    return (ar_ip_addr_and_bearer_token_list)

def update_ntp_worksheet(received_ar_ip_addr,received_ntp_config_list,received_workbook,received_ntp_worksheet_name):
    
    ntp_work_sheet = received_workbook[received_ntp_worksheet_name]
       
    for ntp_data in received_ntp_config_list['items']:
        row_num=ntp_work_sheet.max_row+1
        ntp_work_sheet.cell(row_num,1).value=received_ar_ip_addr
        ntp_work_sheet.cell(row_num,2).value=ntp_data['server_id']
        ntp_work_sheet.cell(row_num,3).value=ntp_data['address']
        ntp_work_sheet.cell(row_num,4).value=ntp_data['prefer']
        ntp_work_sheet.cell(row_num,5).value=ntp_data['version']
        ntp_work_sheet.cell(row_num,6).value=ntp_data['encryption']
        #if there is encryption add to worksheet
        if (ntp_work_sheet.cell(row_num,6).value == 'md5' or 
            ntp_work_sheet.cell(row_num,6).value =='sha1'):

            ntp_work_sheet.cell(row_num,7).value=ntp_data['key_id']
            ntp_work_sheet.cell(row_num,8).value=ntp_data['secret']

def update_snmp_worksheet(received_ar_ip_addr,received_snmp_config,received_workbook,received_snmp_worksheet_name):
    
    snmp_work_sheet = received_workbook[received_snmp_worksheet_name]
    row_num = snmp_work_sheet.max_row+1
    snmp_work_sheet.cell(row_num,1).value=received_ar_ip_addr
    snmp_work_sheet.cell(row_num,2).value=received_snmp_config.get('enabled')
    snmp_work_sheet.cell(row_num,3).value=received_snmp_config.get('contact')
    snmp_work_sheet.cell(row_num,4).value=received_snmp_config.get('location')
    snmp_work_sheet.cell(row_num,5).value=received_snmp_config.get('description')
    version_configuration = received_snmp_config['version_configuration']
    snmp_work_sheet.cell(row_num,6).value=version_configuration.get('version')
    snmp_work_sheet.cell(row_num,7).value=version_configuration.get('community_string')
    snmp_work_sheet.cell(row_num,8).value=version_configuration.get('username')
    snmp_work_sheet.cell(row_num,9).value=version_configuration.get('security_model')
    snmp_work_sheet.cell(row_num,10).value=version_configuration.get('auth_protocol')
    snmp_work_sheet.cell(row_num,11).value=version_configuration.get('authentication_passphrase')
    snmp_work_sheet.cell(row_num,12).value=version_configuration.get('privacy_protocol')
    snmp_work_sheet.cell(row_num,13).value=version_configuration.get('privacy_passphrase')

def update_vifgs_worksheet(received_ar_ip_addr,received_vifgs_configs,received_workbook,received_vifgs_worksheet_name):
    
    vifgs_work_sheet = received_workbook[received_vifgs_worksheet_name]

    for vifgs_data in received_vifgs_configs["items"]:
        row_num = vifgs_work_sheet.max_row+1
        vifgs_config = vifgs_data["config"]
        vifgs_work_sheet.cell(row_num,1).value=received_ar_ip_addr
        vifgs_work_sheet.cell(row_num,2).value=vifgs_config["name"]
        vifgs_work_sheet.cell(row_num,3).value=vifgs_data["id"]
        vifgs_work_sheet.cell(row_num,4).value=vifgs_config["enabled"]
        vifgs_work_sheet.cell(row_num,5).value=vifgs_config["dedup"]
        vifgs_work_sheet.cell(row_num,6).value=str(vifgs_config["members"])
        vifgs_work_sheet.cell(row_num,7).value=vifgs_config["description"]
        vifgs_work_sheet.cell(row_num,8).value=vifgs_config["filter"]["type"]
        vifgs_work_sheet.cell(row_num,9).value=vifgs_config["filter"]["value"]
        vifgs_work_sheet.cell(row_num,10).value=vifgs_config["bandwidth_capacity"]
        vifgs_work_sheet.cell(row_num,11).value=vifgs_config["is_other_vifg"]

def update_common_worksheet(received_ar_ip_addr,received_configs,received_workbook,received_worksheet_name):

    common_worksheet = received_workbook[received_worksheet_name]
    row_num = common_worksheet.max_row+1
    common_worksheet.cell(row_num,1).value=received_ar_ip_addr
    common_worksheet.cell(row_num,2).value=received_configs.get("device_name")
    common_worksheet.cell(row_num,3).value=received_configs.get("sw_version")
    common_worksheet.cell(row_num,4).value=received_configs.get("hw_version")
    common_worksheet.cell(row_num,5).value=str(received_configs.get("mgmt_addresses"))
    common_worksheet.cell(row_num,6).value=received_configs.get("serial")
    common_worksheet.cell(row_num,7).value=received_configs.get("model")

def update_dns_worksheet(received_ar_ip_addr,received_configs,received_workbook,received_worksheet_name):
    ntp_worksheet = received_workbook[received_worksheet_name]
    row_num = ntp_worksheet.max_row+1
    ntp_worksheet.cell(row_num,1).value=received_ar_ip_addr 
    ntp_worksheet.cell(row_num,2).value=received_configs.get("hostname")
    ntp_worksheet.cell(row_num,3).value=str(received_configs.get("dns_servers"))
    ntp_worksheet.cell(row_num,4).value=str(received_configs.get("dns_domains"))

def update_hostgroups_worksheet(received_ar_ip_addr,received_configs,received_workbook,received_worksheet_name):
    hg_worksheet = received_workbook[received_worksheet_name]
    row_num = hg_worksheet.max_row+1 
    
    for item in received_configs['items']:
        hg_worksheet.cell(row_num,1).value=received_ar_ip_addr 
        hg_worksheet.cell(row_num,2).value=item.get("name")
        hg_worksheet.cell(row_num,3).value=item.get("description")
        hg_worksheet.cell(row_num,4).value=item.get("enabled")
        hg_worksheet.cell(row_num,5).value=str(item.get("hosts"))

        row_num += 1

def update_urls_worksheet(received_ar_ip_addr,received_configs,received_workbook,received_worksheet_name):
    urls_worksheet = received_workbook[received_worksheet_name]
    row_num = urls_worksheet.max_row+1

    for item in received_configs['items']:
        urls_worksheet.cell(row_num,1).value=received_ar_ip_addr 
        urls_worksheet.cell(row_num,2).value=item.get("name")
        urls_worksheet.cell(row_num,3).value=item.get("desc")
        urls_worksheet.cell(row_num,4).value=item.get("enabled")
        urls_worksheet.cell(row_num,5).value=item.get("preferred")
        urls_worksheet.cell(row_num,6).value=str(item.get("urls"))

def update_apps_worksheet(received_ar_ip_addr,received_configs,received_workbook,received_worksheet_name):
    apps_worksheet = received_workbook[received_worksheet_name]
    row_num = apps_worksheet.max_row+1

    for item in received_configs['items']:
        apps_worksheet.cell(row_num,1).value=received_ar_ip_addr 
        apps_worksheet.cell(row_num,2).value=item.get("name")
        apps_worksheet.cell(row_num,3).value=item.get("enabled")
        apps_worksheet.cell(row_num,4).value=item.get("traffic_match_mode")
        apps_worksheet.cell(row_num,5).value=item.get("include_dpi_tags")
        definitions = item['definitions']
        for def_item in definitions['items']:
            apps_worksheet.cell(row_num,6).value=str(def_item.get("hosts"))
            for trans_rules in def_item['transport_rules']:
                apps_worksheet.cell(row_num,7).value=trans_rules.get("ports")
                apps_worksheet.cell(row_num,8).value=trans_rules.get("ip_protocol")
                row_num += 1

        row_num += 1

def update_phy_int_worksheet(received_ar_ip_addr,received_configs,received_workbook,received_worksheet_name):
    phy_int_worksheet = received_workbook[received_worksheet_name]
    row_num = phy_int_worksheet.max_row+1
    for item in received_configs["items"]:
        config = item["config"]
        phy_int_worksheet.cell(row_num,1).value=received_ar_ip_addr 
        phy_int_worksheet.cell(row_num,2).value=config.get("alias")
        phy_int_worksheet.cell(row_num,3).value=config.get("enabled")
        phy_int_worksheet.cell(row_num,4).value=config.get("description")
        phy_int_worksheet.cell(row_num,5).value=config.get("speed_duplex")
        state = item["state"]
        phy_int_worksheet.cell(row_num,6).value=state.get("interface_type")
        phy_int_worksheet.cell(row_num,7).value=state.get("status")
        phy_int_worksheet.cell(row_num,8).value=state.get("mtu")

        row_num += 1

def update_cap_jobs_worksheet(received_ar_ip_addr,received_configs,received_workbook,received_worksheet_name):
    cap_jobs_worksheet = received_workbook[received_worksheet_name]
    row_num = cap_jobs_worksheet.max_row+1
    for item in received_configs["items"]:
        config = item["config"]
        cap_jobs_worksheet.cell(row_num,1).value=received_ar_ip_addr 
        cap_jobs_worksheet.cell(row_num,2).value=config.get("name")
        cap_jobs_worksheet.cell(row_num,3).value=str(config.get("vifgs"))
        cap_jobs_worksheet.cell(row_num,4).value=config.get("capture_from_all_vifgs")
        cap_jobs_worksheet.cell(row_num,5).value=config.get("snap_len")
        cap_jobs_worksheet.cell(row_num,6).value=config.get("enabled")
        index = config['indexing']

        cap_jobs_worksheet.cell(row_num,7).value=index.get("enabled")
        retention_rules = index['retention_rules']
        cap_jobs_worksheet.cell(row_num,8).value=retention_rules.get("min_disk_space")
        cap_jobs_worksheet.cell(row_num,9).value=retention_rules.get("max_disk_space")
        cap_jobs_worksheet.cell(row_num,10).value=retention_rules.get("min_retention_time")
        cap_jobs_worksheet.cell(row_num,11).value=retention_rules.get("max_retention_time")
        packet_data_retention_rules = config["retention_rules"]
        cap_jobs_worksheet.cell(row_num,12).value=config.get("optimize_for_read")
        cap_jobs_worksheet.cell(row_num,13).value=packet_data_retention_rules.get("min_disk_space")
        cap_jobs_worksheet.cell(row_num,14).value=packet_data_retention_rules.get("max_disk_space")
        cap_jobs_worksheet.cell(row_num,15).value=packet_data_retention_rules.get("min_retention_time")
        cap_jobs_worksheet.cell(row_num,16).value=packet_data_retention_rules.get("max_retention_time")

        row_num += 1

def get_AR_configuration_and_update_AR_mgmt_spreadsheet(received_username,received_password,received_csv_file_path):

    orginal_ar_list = get_ar_list_from_ar_list_csv_file(received_csv_file_path)

    #if list isn't empty
    if orginal_ar_list:

        active_ar_list=confirm_rest_api_access_to_ARs_and_get_bearer_token(received_username,received_password,orginal_ar_list)

        #will setup spreadsheet and all relevant worksheets
        setup_management_spreadsheet(MGMT_SPREADSHEET_PATH)

        ars = []
        for ar_info in active_ar_list:

            Ar_Device = appresponse_device.AppResponse(ar_info['ar_ip_addr'],ar_info['bear_token'])
 
            update_spreadsheet_with_config(MGMT_SPREADSHEET_PATH,Ar_Device)

            ars.append(Ar_Device)
    else:
        print(f"unable to list of AppResponses from -----> {received_csv_file_path}")

def get_ar_list_from_ar_list_csv_file(received_csv_file_path):
    ar_list = []

    try:
        with open(received_csv_file_path,'r') as csv_file:
            #read data in from csv into a dictionary. Import that header information has a header called "ar_list"
            csv_reader = csv.DictReader(csv_file)

            for line in csv_reader:
                ar_list.append(line.get('ar_list'))
    except:
        print(f"faile to find---> {AR_LIST_CSV_FILE_PATH}")

    #if returned import is "[None]" then we convert it to an empty list
    if ar_list == [None]:
        ar_list = []

    return ar_list

def  main(received_username,received_password):
    pass

if __name__ == "__main__":
    ap = argparse.ArgumentParser()
    ap.add_argument("-u","--username", required=True,help="username")
    args = vars(ap.parse_args())
    username = format(args["username"])
    password = getpass()
    get_AR_configuration_and_update_AR_mgmt_spreadsheet(username,password,AR_LIST_CSV_FILE_PATH)